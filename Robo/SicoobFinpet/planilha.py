from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl import Workbook
from datetime import datetime


HEADERS = [
    "Data Sicoob",
    "Data Simplesvet",
    "Conciliado",
    "Valor Sicoob (R$)",
    "Valor Simplesvet (R$)",
    "Descrição Sicoob",
    "Info Complementar",
    "Bandeira",
]

LARGURAS = [12, 12, 12, 16, 18, 35, 50, 18]

BORDA = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

ESTILOS = {
    "header_font": Font(bold=True, size=11, color="FFFFFF"),
    "header_fill": PatternFill("solid", fgColor="2E7D32"),
    "cell_font": Font(size=10, color="212121"),
    "zebra_fill": PatternFill("solid", fgColor="E8F5E9"),
    "branco_fill": PatternFill("solid", fgColor="FFFFFF"),
    "sim_fill": PatternFill("solid", fgColor="C8E6C9"),
    "nao_fill": PatternFill("solid", fgColor="FFCDD2"),
    "total_font": Font(bold=True, size=11, color="FFFFFF"),
    "total_fill": PatternFill("solid", fgColor="37474F"),
    "centro": Alignment(horizontal="center", vertical="center"),
}


def aplicar_header(ws):
    for col, nome in enumerate(HEADERS, 1):
        cel = ws.cell(row=1, column=col, value=nome)
        cel.font = ESTILOS["header_font"]
        cel.fill = ESTILOS["header_fill"]
        cel.alignment = ESTILOS["centro"]
        cel.border = BORDA


def aplicar_linha(ws, linha, conciliado, zebra=False):
    for col in range(1, len(HEADERS) + 1):
        cel = ws.cell(row=linha, column=col)
        cel.font = ESTILOS["cell_font"]
        cel.alignment = ESTILOS["centro"]
        cel.border = BORDA

        if col == 3:
            cel.fill = ESTILOS["sim_fill"] if conciliado else ESTILOS["nao_fill"]
        else:
            cel.fill = ESTILOS["zebra_fill"] if zebra else ESTILOS["branco_fill"]

        if col in [4, 5]:
            cel.number_format = "R$ #,##0.00"


def aplicar_totais(ws, linha, total_sicoob, total_erp):
    for col in range(1, len(HEADERS) + 1):
        cel = ws.cell(row=linha, column=col)
        cel.font = ESTILOS["total_font"]
        cel.fill = ESTILOS["total_fill"]
        cel.alignment = ESTILOS["centro"]
        cel.border = BORDA

    ws.cell(row=linha, column=1, value="TOTAL")
    ws.cell(row=linha, column=4, value=total_sicoob).number_format = "R$ #,##0.00"
    ws.cell(row=linha, column=5, value=total_erp).number_format = "R$ #,##0.00"


def aplicar_larguras(ws):
    for col, largura in enumerate(LARGURAS, 1):
        ws.column_dimensions[get_column_letter(col)].width = largura


def extrair_valores(item):
    return [
        item.get("data_sicoob") or "",
        item.get("data_erp") or "",
        "SIM" if item.get("conciliado") else "NÃO",
        item.get("valor_sicoob") or 0,
        item.get("valor_erp") or 0,
        item.get("descricao_sicoob") or "",
        item.get("info_complementar") or "",
        item.get("bandeira") or "",
    ]


def parse_data(valor):
    if not valor:
        return None
    for fmt in ["%Y-%m-%d", "%d/%m/%Y"]:
        try:
            return datetime.strptime(valor, fmt)
        except ValueError:
            continue
    return None


def agrupar_por_mes(itens):
    grupos = defaultdict(list)
    for item in itens:
        data = parse_data(item.get("data_sicoob", ""))
        if data:
            chave = f"{data.month:02d}-{data.year}"
            grupos[chave].append(item)
    return grupos


def ordenar_por_data(itens):
    return sorted(
        itens,
        key=lambda x: parse_data(x.get("data_sicoob", "")) or datetime.min,
        reverse=True,
    )


def ordenar_chaves_mes(chaves):
    def parse_chave(chave):
        mes, ano = chave.split("-")
        return (int(ano), int(mes))

    return sorted(chaves, key=parse_chave)


def criar_folha(wb, nome, itens):
    ws = wb.create_sheet(title=nome)
    aplicar_header(ws)

    itens_ordenados = ordenar_por_data(itens)
    total_sicoob = 0
    total_erp = 0

    for i, item in enumerate(itens_ordenados, 2):
        valores = extrair_valores(item)
        for col, valor in enumerate(valores, 1):
            ws.cell(row=i, column=col, value=valor)
        aplicar_linha(ws, i, item.get("conciliado", False), zebra=(i % 2 == 0))

        total_sicoob += item.get("valor_sicoob") or 0
        total_erp += item.get("valor_erp") or 0

    if itens_ordenados:
        linha_total = len(itens_ordenados) + 2
        aplicar_totais(ws, linha_total, total_sicoob, total_erp)

    aplicar_larguras(ws)
    ws.freeze_panes = "A2"

    if itens_ordenados:
        ultima_linha = len(itens_ordenados) + 1
        ultima_coluna = get_column_letter(len(HEADERS))
        ws.auto_filter.ref = f"A1:{ultima_coluna}{ultima_linha}"


def criar_planilha(itens, caminho):
    wb = Workbook()

    grupos = agrupar_por_mes(itens)
    chaves_ordenadas = ordenar_chaves_mes(list(grupos.keys()))

    for chave in chaves_ordenadas:
        criar_folha(wb, chave, grupos[chave])

    # Só remove a planilha padrão se outras planilhas foram criadas
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(caminho)
    return caminho


def filtrar_periodo(itens, mes, ano):
    resultado = []
    for item in itens:
        data = parse_data(item.get("data_sicoob", ""))
        if data and data.month == mes and data.year == ano:
            resultado.append(item)
    return resultado
