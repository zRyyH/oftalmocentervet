from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from utils import parse_data, agrupar_por_mes, ordenar_por_data, ordenar_chaves_mes
import unicodedata


HEADERS = [
    "Data Sicoob",
    "Data Simplesvet",
    "Conciliado",
    "Valor Sicoob (R$)",
    "Valor Simplesvet (R$)",
    "Forma Pag. ERP",
    "Descrição Sicoob",
    "Descrição Simplesvet",
    "Fornecedor Simplesvet",
    "Info Complementar",
]

DESCRICAO_FATURA = "DEB.CONV.DEMAIS EMPRESAS"
TEXTO_CARTAO = "Apenas no relatório do cartão"


def normalizar_texto(texto):
    """Remove acentos e normaliza para comparação segura."""
    texto = unicodedata.normalize("NFKD", str(texto))
    return "".join(c for c in texto if not unicodedata.combining(c)).upper().strip()


def criar_estilos():
    borda = Side(style="thin", color="808080")
    return {
        "header_font": Font(bold=True, color="000000", size=11),
        "header_fill": PatternFill("solid", fgColor="90EE90"),
        "cell_font": Font(size=10),
        "cell_font_erro": Font(size=10, color="CC0000"),
        "alt_fill": PatternFill("solid", fgColor="E8F5E9"),
        "sim_fill": PatternFill("solid", fgColor="C6EFCE"),
        "nao_fill": PatternFill("solid", fgColor="FFC7CE"),
        "fatura_fill": PatternFill("solid", fgColor="ADD8E6"),
        "estorno_vinculado_fill": PatternFill("solid", fgColor="FFE4B5"),
        "estorno_sem_par_fill": PatternFill("solid", fgColor="FFCCCC"),
        "borda": Border(left=borda, right=borda, top=borda, bottom=borda),
        "centro": Alignment(horizontal="center", vertical="center"),
        "esquerda": Alignment(horizontal="left", vertical="center"),
    }


def eh_fatura(item):
    desc = normalizar_texto(item.get("descricao_sicoob") or "")
    return desc == DESCRICAO_FATURA


def get_alinhamento(col, estilos):
    colunas_esquerda = [7, 8, 9, 10]
    return estilos["esquerda"] if col in colunas_esquerda else estilos["centro"]


def aplicar_header(ws, estilos):
    for col, nome in enumerate(HEADERS, 1):
        celula = ws.cell(row=1, column=col, value=nome)
        celula.font = estilos["header_font"]
        celula.fill = estilos["header_fill"]
        celula.alignment = estilos["centro"]
        celula.border = estilos["borda"]


def aplicar_linha(ws, linha, estilos, item, zebra=False):
    conciliado = item.get("conciliado", False)
    match = item.get("match", True)
    is_fatura = eh_fatura(item)
    estorno_vinculado = item.get("estorno_vinculado", False)
    estorno_sem_par = item.get("estorno_sem_par", False)
    forma_confere = item.get("forma_confere")

    for col in range(1, len(HEADERS) + 1):
        celula = ws.cell(row=linha, column=col)
        celula.border = estilos["borda"]
        celula.alignment = get_alinhamento(col, estilos)

        # Coluna 6 é "Forma Pag. ERP" - vermelha se forma_confere é False
        if col == 6 and forma_confere is False:
            celula.font = estilos["cell_font_erro"]
        elif match is False:
            celula.font = estilos["cell_font_erro"]
        else:
            celula.font = estilos["cell_font"]

        # Prioridade: estorno_sem_par > estorno_vinculado > fatura > conciliado > zebra
        if estorno_sem_par:
            celula.fill = estilos["estorno_sem_par_fill"]
        elif estorno_vinculado:
            celula.fill = estilos["estorno_vinculado_fill"]
        elif is_fatura:
            celula.fill = estilos["fatura_fill"]
        elif col == 3 and not estorno_vinculado:
            # Não aplica cor verde/vermelho na coluna 3 se for estorno vinculado
            celula.fill = estilos["sim_fill"] if conciliado else estilos["nao_fill"]
        elif zebra:
            celula.fill = estilos["alt_fill"]


def auto_ajustar_larguras(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2


def extrair_valores(item):
    is_fatura = eh_fatura(item)
    estorno_vinculado = item.get("estorno_vinculado", False)

    if is_fatura:
        return [
            item.get("data_sicoob", ""),
            item.get("data_erp") or "",
            "FATURA",
            item.get("valor_sicoob"),
            item.get("valor_erp"),
            item.get("forma_pagamento_erp") or "",
            TEXTO_CARTAO,
            TEXTO_CARTAO,
            TEXTO_CARTAO,
            TEXTO_CARTAO,
        ]

    # Se é estorno vinculado, campo conciliado fica em branco
    if estorno_vinculado:
        status_conciliado = ""
    else:
        status_conciliado = "SIM" if item.get("conciliado") else "NÃO"

    return [
        item.get("data_sicoob", ""),
        item.get("data_erp") or "",
        status_conciliado,
        item.get("valor_sicoob"),
        item.get("valor_erp"),
        item.get("forma_pagamento_erp") or "",
        item.get("descricao_sicoob", ""),
        item.get("descricao_erp") or "",
        item.get("fornecedor_erp") or "",
        item.get("info_complementar", ""),
    ]




def criar_folha(wb, nome, itens, estilos):
    ws = wb.create_sheet(title=nome)
    aplicar_header(ws, estilos)

    itens = ordenar_por_data(itens, campo_data="data_sicoob")

    for i, item in enumerate(itens, 2):
        valores = extrair_valores(item)
        for col, valor in enumerate(valores, 1):
            ws.cell(row=i, column=col, value=valor)
        aplicar_linha(ws, i, estilos, item, zebra=(i % 2 == 0))

    auto_ajustar_larguras(ws)
    ws.freeze_panes = "A2"

    if itens:
        ultima_linha = len(itens) + 1
        ultima_coluna = get_column_letter(len(HEADERS))
        ws.auto_filter.ref = f"A1:{ultima_coluna}{ultima_linha}"


def criar_planilha(itens, caminho):
    wb = Workbook()
    estilos = criar_estilos()

    grupos = agrupar_por_mes(itens, campo_data="data_sicoob")
    chaves_ordenadas = ordenar_chaves_mes(grupos.keys())

    for chave in chaves_ordenadas:
        criar_folha(wb, chave, grupos[chave], estilos)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(caminho)
    return caminho