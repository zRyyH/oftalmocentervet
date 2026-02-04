from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl import Workbook
from datetime import datetime

BORDA = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)

CORES = {
    "header": PatternFill("solid", fgColor="2E7D32"),
    "zebra": PatternFill("solid", fgColor="E8F5E9"),
    "branco": PatternFill("solid", fgColor="FFFFFF"),
    "sim": PatternFill("solid", fgColor="C8E6C9"),
    "nao": PatternFill("solid", fgColor="FFCDD2"),
    "vazio": PatternFill("solid", fgColor="FFCCCC"),
    "total": PatternFill("solid", fgColor="37474F"),
    "destaque": PatternFill("solid", fgColor="ADD8E6"),
    "estorno": PatternFill("solid", fgColor="FFE0B2"),  # Laranja suave
    "devolucao": PatternFill("solid", fgColor="E1BEE7"),  # Roxo suave
    "resumo_header": PatternFill("solid", fgColor="1565C0"),
    "resumo_azul": PatternFill("solid", fgColor="E3F2FD"),
    "resumo_laranja": PatternFill("solid", fgColor="FFF3E0"),
    "resumo_verde_header": PatternFill("solid", fgColor="2E7D32"),
    "resumo_verde": PatternFill("solid", fgColor="E8F5E9"),
}

FONTES = {
    "header": Font(bold=True, size=11, color="FFFFFF"),
    "celula": Font(size=10),
    "erro": Font(size=10, color="CC0000"),
    "aviso": Font(size=10, color="0000CC"),
    "total": Font(bold=True, size=11, color="FFFFFF"),
    "resumo_azul": Font(bold=True, size=10, color="1565C0"),
    "resumo_laranja": Font(bold=True, size=10, color="E65100"),
    "resumo_verde": Font(bold=True, size=10, color="2E7D32"),
    "resumo_valor": Font(bold=True, size=10, color="000000"),
}


def _parse_data(valor):
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(valor[:10], fmt)
            except:
                pass
    return None


def _detectar_coluna_data(campos):
    for c in campos:
        if "data" in c.lower() or c.lower() in ("date", "dt", "created_at"):
            return c
    return None


def _aplicar_header(ws, headers, config):
    for col, nome in enumerate(headers, 1):
        cel = ws.cell(row=1, column=col, value=nome)
        cel.font = FONTES["header"]
        cel.fill = CORES["header"]
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = BORDA


def _aplicar_linha(ws, linha, valores, config, zebra=False):
    colunas_status = config.get("colunas_status", [])
    colunas_moeda = config.get("colunas_moeda", [])
    colunas_erro = config.get("colunas_erro", {})
    colunas_destaque = config.get("colunas_destaque", [])
    marcar_vazios = config.get("marcar_vazios", False)
    cor_linha = config.get("cor_linha")  # Cor para toda a linha

    for col, valor in enumerate(valores, 1):
        cel = ws.cell(row=linha, column=col, value=valor)
        cel.border = BORDA
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.font = FONTES["celula"]

        # Cor de fundo base - prioridade: cor_linha > destaque > status > zebra
        if cor_linha and cor_linha in CORES:
            cel.fill = CORES[cor_linha]
        elif col in colunas_destaque:
            cel.fill = CORES["destaque"]
        elif col in colunas_status:
            if valor == "SIM":
                cel.fill = CORES["sim"]
            elif valor in ("NÃO", "NAO"):
                cel.fill = CORES["nao"]
            else:
                cel.fill = CORES["zebra"] if zebra else CORES["branco"]
        elif marcar_vazios and (valor is None or valor == ""):
            cel.fill = CORES["vazio"]
        else:
            cel.fill = CORES["zebra"] if zebra else CORES["branco"]

        # Formatação moeda
        if col in colunas_moeda:
            cel.number_format = "R$ #,##0.00"

        # Fonte de erro/aviso
        if col in colunas_erro:
            tipo = colunas_erro[col]
            if tipo == "erro":
                cel.font = FONTES["erro"]
            elif tipo == "aviso":
                cel.font = FONTES["aviso"]


def _aplicar_totais(ws, linha, totais, config):
    colunas_moeda = config.get("colunas_moeda", [])

    for col in range(1, ws.max_column + 1):
        cel = ws.cell(row=linha, column=col)
        cel.font = FONTES["total"]
        cel.fill = CORES["total"]
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = BORDA

    ws.cell(row=linha, column=1, value="TOTAL")

    for col, valor in totais.items():
        cel = ws.cell(row=linha, column=col, value=valor)
        if col in colunas_moeda:
            cel.number_format = "R$ #,##0.00"


def _aplicar_resumo(ws, linha_inicio, dados, config):
    """
    Aplica área de resumo no final da folha.

    config["resumo"] deve conter:
        - titulo: título do resumo (ex: "RESUMO GERAL")
        - cor_header: cor do cabeçalho ("azul", "verde", etc.)
        - linhas: lista de dicts com:
            - label: texto do label
            - tipo: "contagem", "soma", "quantidade" ou "total_registros"
            - campo: campo para somar (se tipo="soma" ou "quantidade")
            - filtro: dict {campo: valor} para filtrar registros
            - cor: "azul", "laranja" ou "verde" (default: "azul")
    """
    resumo_config = config.get("resumo")
    if not resumo_config:
        return linha_inicio

    titulo = resumo_config.get("titulo", "RESUMO GERAL")
    linhas = resumo_config.get("linhas", [])
    cor_header = resumo_config.get("cor_header", "azul")
    num_colunas = ws.max_column

    # Linha em branco para separação
    linha_inicio += 2

    # Cabeçalho do resumo
    header_fill = CORES.get(f"resumo_{cor_header}_header", CORES["resumo_header"])
    for col in range(1, num_colunas + 1):
        cel = ws.cell(row=linha_inicio, column=col)
        cel.font = FONTES["header"]
        cel.fill = header_fill
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = BORDA

    ws.cell(row=linha_inicio, column=1, value=titulo)
    ws.merge_cells(f"A{linha_inicio}:{get_column_letter(num_colunas)}{linha_inicio}")

    # Linhas do resumo
    for item_resumo in linhas:
        linha_inicio += 1
        label = item_resumo.get("label", "")
        tipo = item_resumo.get("tipo", "contagem")
        campo = item_resumo.get("campo")
        filtro = item_resumo.get("filtro", {})
        cor = item_resumo.get("cor", "azul")

        # Aplica estilo nas células
        cor_fill = CORES.get(f"resumo_{cor}", CORES["resumo_azul"])
        cor_font = FONTES.get(f"resumo_{cor}", FONTES["resumo_azul"])

        for col in range(1, num_colunas + 1):
            cel = ws.cell(row=linha_inicio, column=col)
            cel.border = BORDA
            if col <= 4:
                cel.fill = cor_fill
            else:
                cel.fill = CORES["branco"]

        # Label
        ws.cell(row=linha_inicio, column=1, value=label)
        ws.cell(row=linha_inicio, column=1).font = cor_font
        ws.cell(row=linha_inicio, column=1).alignment = Alignment(
            horizontal="left", vertical="center"
        )
        ws.merge_cells(f"A{linha_inicio}:C{linha_inicio}")

        # Calcula valor baseado no tipo e filtro
        dados_filtrados = dados
        if filtro:
            dados_filtrados = [
                d for d in dados
                if all(d.get(k) == v for k, v in filtro.items())
            ]

        if tipo == "total_registros":
            valor = len(dados)
        elif tipo == "contagem":
            valor = len(dados_filtrados)
        elif tipo in ("soma", "quantidade") and campo:
            valor = sum(
                d.get(campo, 0) or 0
                for d in dados_filtrados
                if isinstance(d.get(campo), (int, float))
            )
        else:
            valor = 0

        cel_valor = ws.cell(row=linha_inicio, column=4, value=valor)
        cel_valor.font = FONTES["resumo_valor"]
        cel_valor.alignment = Alignment(horizontal="center", vertical="center")

        if tipo == "soma":
            cel_valor.number_format = "R$ #,##0.00"

    return linha_inicio


def _ajustar_larguras(ws, larguras=None):
    for col_idx, col in enumerate(ws.columns, 1):
        if larguras and col_idx <= len(larguras):
            ws.column_dimensions[get_column_letter(col_idx)].width = larguras[
                col_idx - 1
            ]
        else:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 2, 50
            )


def _criar_folha(wb, nome, dados, headers, campos, config):
    ws = wb.create_sheet(title=nome[:31])
    _aplicar_header(ws, headers, config)

    colunas_soma = config.get("colunas_soma", [])
    totais = {col: 0 for col in colunas_soma}

    for i, item in enumerate(dados, 2):
        valores = [item.get(c) for c in campos]

        # Monta config dinâmica para linha
        linha_config = config.copy()

        # Suporte a erros por linha
        if "_erros" in item:
            linha_config["colunas_erro"] = {
                campos.index(k) + 1: v for k, v in item["_erros"].items() if k in campos
            }

        # Suporte a destaque por linha
        if "_destaque" in item:
            linha_config["colunas_destaque"] = [
                campos.index(k) + 1 for k in item["_destaque"] if k in campos
            ]

        # Suporte a cor de linha inteira
        if "_cor_linha" in item:
            linha_config["cor_linha"] = item["_cor_linha"]

        _aplicar_linha(ws, i, valores, linha_config, zebra=(i % 2 == 0))

        # Acumula totais
        for col in colunas_soma:
            val = item.get(campos[col - 1])
            if isinstance(val, (int, float)):
                totais[col] += val

    # Linha de totais
    linha_atual = len(dados) + 1
    if colunas_soma and dados:
        linha_atual = len(dados) + 2
        _aplicar_totais(ws, linha_atual, totais, config)

    # Área de resumo
    if config.get("resumo") and dados:
        _aplicar_resumo(ws, linha_atual, dados, config)

    _ajustar_larguras(ws, config.get("larguras"))
    ws.freeze_panes = "A2"

    ultima_linha = len(dados) + 1
    ultima_coluna = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{ultima_coluna}{ultima_linha}"


def criar_planilha(dados, arquivo, headers=None, campos=None, config=None):
    """
    Cria planilha Excel com dados agrupados por mês.

    config aceita:
        - coluna_data: campo para agrupar por mês (auto-detecta se None)
        - colunas_status: lista de índices (1-based) para colorir SIM/NÃO
        - colunas_moeda: lista de índices (1-based) para formato R$
        - colunas_soma: lista de índices (1-based) para somar e exibir total
        - larguras: lista de larguras por coluna
        - marcar_vazios: True para pintar células vazias de rosa
        - resumo: dict para área de resumo no final de cada folha:
            - titulo: título do resumo (ex: "RESUMO GERAL")
            - cor_header: cor do cabeçalho ("azul", "verde", etc.)
            - linhas: lista de dicts com:
                - label: texto do label
                - tipo: "contagem", "soma", "quantidade" ou "total_registros"
                - campo: campo para somar (se tipo="soma" ou "quantidade")
                - filtro: dict {campo: valor} para filtrar registros
                - cor: "azul", "laranja" ou "verde" (default: "azul")

    Cada item pode ter:
        - _erros: dict {campo: "erro"|"aviso"} para fonte colorida
        - _destaque: lista de campos para fundo azul claro
        - _cor_linha: nome da cor para aplicar em toda a linha (ex: "estorno")
    """
    if not dados:
        return None

    config = config or {}

    # Detecta campos automaticamente
    campos_internos = {"_erros", "_destaque"}
    if campos is None:
        campos = [k for k in dados[0].keys() if k not in campos_internos]
    if headers is None:
        headers = campos

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    coluna_data = config.get("coluna_data") or _detectar_coluna_data(campos)

    if coluna_data and coluna_data in campos:
        # Agrupa por mês
        grupos = defaultdict(list)
        for item in dados:
            dt = _parse_data(item.get(coluna_data))
            chave = (dt.year, dt.month) if dt else (9999, 99)
            grupos[chave].append((dt or datetime.max, item))

        # Nomes dos meses em português
        meses_pt = {
            1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
            5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
            9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
        }

        for chave in sorted(grupos.keys()):
            itens = sorted(grupos[chave], key=lambda x: x[0], reverse=True)
            dados_ordenados = [item[1] for item in itens]

            if chave == (9999, 99):
                nome = "Sem Data"
            else:
                nome = f"{meses_pt[chave[1]]} {chave[0]}"

            _criar_folha(wb, nome, dados_ordenados, headers, campos, config)
    else:
        _criar_folha(wb, "Dados", dados, headers, campos, config)

    wb.save(arquivo)
    return arquivo
