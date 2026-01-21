from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime


def criar_planilha(resultado: list, caminho: str = "finpet_conciliacoes.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    dados_por_mes = defaultdict(list)
    for item in resultado:
        data = item.get("data", "")
        if isinstance(data, datetime):
            chave = data.strftime("%m/%Y")
        elif isinstance(data, str) and data:
            try:
                dt = datetime.strptime(data, "%d/%m/%Y")
                chave = dt.strftime("%m/%Y")
            except ValueError:
                try:
                    dt = datetime.strptime(data, "%Y-%m-%d")
                    chave = dt.strftime("%m/%Y")
                except ValueError:
                    chave = "Sem Data"
        else:
            chave = "Sem Data"
        dados_por_mes[chave].append(item)

    headers = [
        "Data",
        "Bandeira",
        "Tipo",
        "Valor ERP",
        "Valor Finpet",
        "Diferença",
        "Taxa Bandeira",
        "Taxa Aluguel",
        "Transações Finpet",
        "Conciliado ERP",
        "Bateu",
    ]

    borda_cinza = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )
    header_font = Font(bold=True)
    linha_impar = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    linha_par = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    rosa_vazio = PatternFill(
        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
    )
    fonte_vermelha = Font(color="CC0000")
    verde_suave = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    vermelho_suave = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    meses_ordenados = sorted(
        dados_por_mes.keys(),
        key=lambda x: (
            datetime.strptime(x, "%m/%Y") if x != "Sem Data" else datetime.min
        ),
    )

    for nome_mes in meses_ordenados:
        itens_mes = dados_por_mes[nome_mes]
        ws = wb.create_sheet(title=nome_mes.replace("/", "-"))

        for col, header in enumerate(headers, 1):
            celula = ws.cell(row=1, column=col, value=header)
            celula.fill = header_fill
            celula.font = header_font
            celula.alignment = alinhamento_centro
            celula.border = borda_cinza

        for row_idx, item in enumerate(itens_mes, 2):
            match = item.get("match", True)
            bateu = "SIM" if item.get("bateu") else "NÃO"
            conciliado = item.get("conciliado_erp", "")

            valores = [
                item.get("data", ""),
                item.get("bandeira", ""),
                item.get("tipo", ""),
                item.get("valor_erp"),
                item.get("valor_finpet"),
                item.get("diferenca"),
                item.get("taxa_bandeira"),
                item.get("taxa_aluguel"),
                item.get("transacoes_finpet"),
                conciliado,
                bateu,
            ]
            fill_linha = linha_par if row_idx % 2 == 0 else linha_impar

            for col_idx, valor in enumerate(valores, 1):
                celula = ws.cell(row=row_idx, column=col_idx, value=valor)
                celula.alignment = alinhamento_centro
                celula.border = borda_cinza

                if valor == "SIM":
                    celula.fill = verde_suave
                elif valor == "NÃO":
                    celula.fill = vermelho_suave
                elif valor is None or valor == "":
                    celula.fill = rosa_vazio
                else:
                    celula.fill = fill_linha

                if not match:
                    celula.font = fonte_vermelha

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    wb.save(caminho)
    return caminho
