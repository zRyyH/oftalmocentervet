import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

VERDE_LINHA = PatternFill("solid", fgColor="C6EFCE")
VERMELHO_LINHA = PatternFill("solid", fgColor="FFC7CE")
VERDE_HEADER = PatternFill("solid", fgColor="70AD47")
BORDA = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)
CENTRO = Alignment(horizontal="center", vertical="center")

HEADERS = [
    "Valor Finpet",
    "Data Estimada",
    "Bandeira",
    "Score",
    "Parcela Finpet",
    "Parcela Lançamento",
    "Valor Comparado Finpet",
    "Valor Comparado Lançamento",
    "Data Finpet",
    "Data Lançamento",
    "Autorização Finpet",
    "Autorização Lançamento",
    "Cliente Finpet",
    "Cliente Lançamento",
]


def carregar_dados(caminho):
    with open(caminho, encoding="utf-8") as f:
        conteudo = f.read()
    inicio = conteudo.find("[")
    json_str = conteudo[inicio:].replace("[...", "[")
    return json.loads(json_str)


def formatar_data(valor):
    if not valor:
        return ""
    return valor[:10] if isinstance(valor, str) else ""


def extrair_linha(item):
    fp = item.get("finpet", {})
    score = item.get("score", 0)
    comp = item.get("comparacoes", {})

    cliente_fp = comp.get("cliente", {}).get("finpet", "")
    if isinstance(cliente_fp, list):
        cliente_fp = ", ".join(str(x) for x in cliente_fp)

    return [
        fp.get("value"),
        formatar_data(fp.get("date_estimated")),
        fp.get("payment_brand"),
        score,
        comp.get("parcela", {}).get("finpet"),
        comp.get("parcela", {}).get("release"),
        comp.get("valor", {}).get("finpet"),
        comp.get("valor", {}).get("release"),
        formatar_data(comp.get("data", {}).get("finpet_estimated")),
        formatar_data(comp.get("data", {}).get("release")),
        comp.get("auth", {}).get("finpet"),
        comp.get("auth", {}).get("release"),
        cliente_fp,
        comp.get("cliente", {}).get("release"),
    ]


def criar_planilha(dados, saida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação"

    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = VERDE_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = CENTRO
        cell.border = BORDA

    for item in dados:
        linha = extrair_linha(item)
        ws.append(linha)

        row_idx = ws.max_row
        cor = VERDE_LINHA if item.get("score", 0) >= 5 else VERMELHO_LINHA
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = cor
            cell.alignment = CENTRO
            cell.border = BORDA

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    wb.save(saida)


if __name__ == "__main__":
    dados = carregar_dados("resultado.json")
    criar_planilha(dados, "conciliacao.xlsx")
