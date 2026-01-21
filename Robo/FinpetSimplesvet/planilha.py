from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from collections import defaultdict


def formatar_data(valor):
    if not valor:
        return None
    if isinstance(valor, str) and len(valor) >= 10:
        return valor[:10]
    return valor


def extrair_mes_ano(valor):
    if not valor:
        return None
    if isinstance(valor, str) and len(valor) >= 7:
        partes = valor[:10].split("-")
        if len(partes) >= 2:
            return f"{partes[1]}/{partes[0]}"
    return None


def aplicar_formatacao(ws, headers, dados_mes):
    header_fill = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )
    row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    vazio_fill = PatternFill(
        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
    )
    fonte_vermelha = Font(color="CC0000", bold=True)
    borda_cinza = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )
    centro = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill

    for item in dados_mes:
        finpet = item.get("finpet", {})
        comparacoes = item.get("comparacoes", {})
        pedidos = comparacoes.get("cliente", {}).get("finpet", [])
        pedidos_str = ", ".join(pedidos) if pedidos else ""

        linha = [
            comparacoes.get("valor", {}).get("finpet"),
            formatar_data(comparacoes.get("data", {}).get("finpet_estimated")),
            finpet.get("payment_brand"),
            item.get("score"),
            comparacoes.get("parcela", {}).get("finpet"),
            comparacoes.get("parcela", {}).get("release"),
            comparacoes.get("valor", {}).get("finpet"),
            comparacoes.get("valor", {}).get("release"),
            formatar_data(comparacoes.get("data", {}).get("finpet_estimated")),
            formatar_data(comparacoes.get("data", {}).get("release")),
            comparacoes.get("auth", {}).get("finpet"),
            comparacoes.get("auth", {}).get("release"),
            pedidos_str,
            comparacoes.get("cliente", {}).get("release"),
        ]
        ws.append(linha)

        row_idx = ws.max_row
        colunas_match = {
            6: comparacoes.get("parcela", {}).get("match", True),
            8: comparacoes.get("valor", {}).get("match", True),
            10: comparacoes.get("data", {}).get("match", True),
            12: comparacoes.get("auth", {}).get("match", True),
            13: comparacoes.get("cliente", {}).get("match", True),
        }

        for col_idx, match in colunas_match.items():
            if not match:
                ws.cell(row=row_idx, column=col_idx).font = fonte_vermelha

        for col_idx in range(1, len(linha) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None or cell.value == "":
                cell.fill = vazio_fill

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                if not cell.fill.start_color.rgb.endswith("FFCCCC"):
                    cell.fill = row_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = centro
            cell.border = borda_cinza

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_len:
                    max_len = cell_len
            except:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def gerar_excel(dados: List[Dict[str, Any]], caminho_saida: str = "resultados.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "Valor Finpet",
        "Data Estimada",
        "Bandeira",
        "Score",
        "Parcela Finpet",
        "Parcela Simplesvet",
        "Valor Finpet",
        "Valor Simplesvet",
        "Data Finpet",
        "Data Simplesvet",
        "Auth Finpet",
        "Auth Simplesvet",
        "Pedidos Extraidos",
        "Cliente Lancamento",
    ]

    dados_por_mes = defaultdict(list)
    for item in dados:
        comparacoes = item.get("comparacoes", {})
        data_str = comparacoes.get("data", {}).get("finpet_estimated")
        mes_ano = extrair_mes_ano(data_str)
        if mes_ano:
            dados_por_mes[mes_ano].append(item)
        else:
            dados_por_mes["Sem Data"].append(item)

    def ordenar_mes(mes_ano):
        if mes_ano == "Sem Data":
            return (9999, 99)
        partes = mes_ano.split("/")
        return (int(partes[1]), int(partes[0]))

    meses_ordenados = sorted(dados_por_mes.keys(), key=ordenar_mes)

    for mes_ano in meses_ordenados:
        nome_folha = mes_ano.replace("/", "-")
        ws = wb.create_sheet(title=nome_folha)
        aplicar_formatacao(ws, headers, dados_por_mes[mes_ano])

    wb.save(caminho_saida)
    return caminho_saida
