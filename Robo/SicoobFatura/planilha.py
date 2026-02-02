import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Adiciona o diretório raiz ao path para importar utils
root_dir = Path(__file__).parent.parent
sys.path.insert(0, str(root_dir))

from utils import extrair_mes_ano, formatar_data as formatar_data_utils


COR_CABECALHO = PatternFill("solid", fgColor="90EE90")
COR_ZEBRA = PatternFill("solid", fgColor="E8F5E9")
COR_STATUS_OK = PatternFill("solid", fgColor="C6EFCE")
COR_STATUS_ERRO = PatternFill("solid", fgColor="FFC7CE")
BORDA = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)

COLUNAS = ["Data", "Descrição", "Fornecedor", "Valor", "Parcela", "Vencimento"]


def formatar_data(data_str):
    """Wrapper para formatar data no formato brasileiro."""
    return formatar_data_utils(data_str, formato_saida="%d/%m/%Y")


def formatar_valor(valor):
    if valor is None:
        return 0
    return abs(float(valor))


def aplicar_estilo_cabecalho(ws):
    for col, titulo in enumerate(COLUNAS, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True, size=11)
        cell.fill = COR_CABECALHO
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDA


def ajustar_larguras(ws):
    for col in range(1, len(COLUNAS) + 1):
        max_len = len(COLUNAS[col - 1])
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)


def aplicar_linha_resumo(ws, linha, label, valor, formato=None, destaque=False, cor_fundo=None):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
    ws.merge_cells(start_row=linha, start_column=4, end_row=linha, end_column=6)
    
    cell_label = ws.cell(row=linha, column=1, value=label)
    cell_label.font = Font(bold=True, size=12 if destaque else 11)
    cell_label.alignment = Alignment(horizontal="right", vertical="center")
    cell_label.border = BORDA
    
    cell_valor = ws.cell(row=linha, column=4, value=valor)
    cell_valor.font = Font(bold=True, size=12 if destaque else 11)
    cell_valor.alignment = Alignment(horizontal="center", vertical="center")
    cell_valor.border = BORDA
    if formato:
        cell_valor.number_format = formato
    
    cor = cor_fundo or COR_ZEBRA
    for col in range(1, 7):
        ws.cell(row=linha, column=col).fill = cor
        ws.cell(row=linha, column=col).border = BORDA


def criar_resumo(ws, linha_inicio, dados_mes, despesas):
    valor_fatura = dados_mes["valor_fatura"]
    total_despesas = dados_mes["total_despesas"]
    diferenca = round(valor_fatura - total_despesas, 2)
    tem_erro = abs(diferenca) > 0.01
    
    num_despesas = len(despesas)
    valores = [d["valor"] for d in despesas if d["valor"]]
    maior_despesa = max(valores) if valores else 0
    menor_despesa = min(valores) if valores else 0
    media_despesa = sum(valores) / len(valores) if valores else 0
    fornecedores = set(d["fornecedor"].strip() for d in despesas if d["fornecedor"].strip())
    
    linha = linha_inicio + 2
    
    # Título
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
    cell_titulo = ws.cell(row=linha, column=1, value="📊 RESUMO DA FATURA")
    cell_titulo.font = Font(bold=True, size=14)
    cell_titulo.fill = COR_CABECALHO
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 7):
        ws.cell(row=linha, column=col).border = BORDA
        ws.cell(row=linha, column=col).fill = COR_CABECALHO
    linha += 1
    
    # Dados
    aplicar_linha_resumo(ws, linha, "Nº de Despesas", num_despesas)
    linha += 1
    aplicar_linha_resumo(ws, linha, "Nº de Fornecedores", len(fornecedores))
    linha += 1
    aplicar_linha_resumo(ws, linha, "Maior Despesa", maior_despesa, "R$ #,##0.00")
    linha += 1
    aplicar_linha_resumo(ws, linha, "Menor Despesa", menor_despesa, "R$ #,##0.00")
    linha += 1
    aplicar_linha_resumo(ws, linha, "Média por Despesa", media_despesa, "R$ #,##0.00")
    linha += 2
    
    # Totais
    aplicar_linha_resumo(ws, linha, "TOTAL DESPESAS", total_despesas, "R$ #,##0.00", True, COR_CABECALHO)
    linha += 1
    aplicar_linha_resumo(ws, linha, "VALOR FATURA", valor_fatura, "R$ #,##0.00", True, COR_CABECALHO)
    linha += 1
    
    cor_diff = COR_STATUS_ERRO if tem_erro else COR_STATUS_OK
    aplicar_linha_resumo(ws, linha, "DIFERENÇA", diferenca, "R$ #,##0.00", True, cor_diff)
    linha += 2
    
    # Status
    status = "✅ FATURA CONFERIDA" if not tem_erro else "⚠️ VERIFICAR DIFERENÇA"
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
    cell_status = ws.cell(row=linha, column=1, value=status)
    cell_status.font = Font(bold=True, size=12, color="006400" if not tem_erro else "CC0000")
    cell_status.alignment = Alignment(horizontal="center", vertical="center")


def gerar_planilha_sicoob(dados, caminho_saida="Relatorios/Cartão De Credito.xlsx"):
    registros_por_mes = defaultdict(
        lambda: {"despesas": [], "valor_fatura": 0, "total_despesas": 0}
    )

    for item in dados:
        data_fatura = item.get("data_fatura", "")
        mes_ano = extrair_mes_ano(data_fatura) or "Sem Data"

        registros_por_mes[mes_ano]["valor_fatura"] = item.get("valor_fatura", 0)
        registros_por_mes[mes_ano]["total_despesas"] = item.get("total_despesas", 0)

        for desp in item.get("despesas", []):
            registro = {
                "data": formatar_data(desp.get("data", data_fatura)),
                "descricao": desp.get("descricao", ""),
                "fornecedor": desp.get("fornecedor", "").replace("Cartão de crédito", ""),
                "valor": formatar_valor(desp.get("valor")),
                "parcela": desp.get("parcela", ""),
                "vencimento": formatar_data(desp.get("vencimento", "")),
                "data_sort": desp.get("data", data_fatura),
            }
            registros_por_mes[mes_ano]["despesas"].append(registro)

    wb = Workbook()
    wb.remove(wb.active)

    for mes_ano in sorted(registros_por_mes.keys(), reverse=True):
        dados_mes = registros_por_mes[mes_ano]
        despesas = dados_mes["despesas"]
        despesas.sort(key=lambda x: x["data_sort"], reverse=True)

        ws = wb.create_sheet(title=mes_ano[:31])
        aplicar_estilo_cabecalho(ws)

        linha_atual = 2
        for reg in despesas:
            valores = [
                reg["data"],
                reg["descricao"],
                reg["fornecedor"],
                reg["valor"],
                reg["parcela"],
                reg["vencimento"],
            ]

            for col, valor in enumerate(valores, 1):
                cell = ws.cell(row=linha_atual, column=col, value=valor)
                cell.border = BORDA
                cell.font = Font(size=10)

                if col in [1, 4, 5, 6]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

                if linha_atual % 2 == 0:
                    cell.fill = COR_ZEBRA

            ws.cell(row=linha_atual, column=4).number_format = "R$ #,##0.00"
            linha_atual += 1

        criar_resumo(ws, linha_atual, dados_mes, despesas)
        ajustar_larguras(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:F{linha_atual - 1}"

    wb.save(caminho_saida)
    return caminho_saida